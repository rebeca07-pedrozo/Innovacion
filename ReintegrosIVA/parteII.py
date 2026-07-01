#primero 
import pandas as pd, numpy as np, re, unicodedata
from datetime import datetime, timezone, timedelta
import os
from google.colab import auth
auth.authenticate_user()
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from datetime import datetime, timezone, timedelta
drive = build('drive', 'v3')    
print("Conectado a Google Drive")


#segundo 
# ===================== AQUÍ PONES TU INFORMACIÓN =====================
FOLDER_ID = ""          # <<<<<< ID de tu carpeta de Drive (donde dejas los .txt del lote)

SEPARADOR       = ","   # separador del txt:  ","   "\t"   "|"   ";"
IVA             = 0.19
SIGNO_CUENTA_2  = 1      # 1 = (Créd−Déb).  Pon -1 si el IVA te llega en Débito
TOLERANCIA      = 1.0

COL_CUENTA  = "Cuenta"
COL_DEBITO  = "Debito"
COL_CREDITO = "Crédito"
COL_TIPO    = "Tipo Comprobante"
COL_DESCTRX = "Descripción transacción"

SUB_REPORTE       = "REPORTE_GENERAL"   # Paso 1
SUB_COMPROBANTES  = "COMPROBANTES"      # Paso 2 (AH.xlsx, BR.xlsx, ...)
SUB_TXT_PROC      = "TXT_PROCESADOS"    # acá se mueven los .txt ya procesados
# =====================================================================
print("Configuración lista")

#tercero#tercero
import os

def listar_archivos(folder_id, extensiones=(".txt", ".csv")):
    """Archivos de la carpeta que terminen en .txt o .csv"""
    q = f"'{folder_id}' in parents and trashed = false"
    res = drive.files().list(q=q, fields="files(id, name, mimeType)", pageSize=1000).execute()
    return [f for f in res.get("files", []) if f["name"].lower().endswith(extensiones)]

def descargar(file_id, destino_local):
    """Descarga un archivo de Drive al disco temporal de Colab"""
    req = drive.files().get_media(fileId=file_id)
    with open(destino_local, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    return destino_local

def obtener_o_crear_subcarpeta(nombre, parent_id):
    """Busca la subcarpeta; si no existe, la crea (no duplica)"""
    q = (f"name = '{nombre}' and '{parent_id}' in parents "
         f"and mimeType = 'application/vnd.google-apps.folder' and trashed = false")
    res = drive.files().list(q=q, fields="files(id)").execute().get("files", [])
    if res:
        return res[0]["id"]
    meta = {"name": nombre, "mimeType": "application/vnd.google-apps.folder", "parents": [parent_id]}
    return drive.files().create(body=meta, fields="id").execute()["id"]

def subir(ruta_local, folder_id):
    """Sube un archivo local a la carpeta de Drive indicada"""
    meta = {"name": os.path.basename(ruta_local), "parents": [folder_id]}
    media = MediaFileUpload(ruta_local, resumable=True)
    drive.files().create(body=meta, media_body=media, fields="id").execute()

def buscar_archivo(nombre, folder_id):
    """Devuelve el archivo {id,name} si existe en la carpeta, o None."""
    q = f"name = '{nombre}' and '{folder_id}' in parents and trashed = false"
    res = drive.files().list(q=q, fields="files(id, name)").execute().get("files", [])
    return res[0] if res else None

def actualizar_o_crear(nombre, folder_id, ruta_local):
    """Si el archivo ya existe en Drive lo ACTUALIZA; si no, lo crea."""
    media = MediaFileUpload(ruta_local, resumable=True)
    ex = buscar_archivo(nombre, folder_id)
    if ex:
        drive.files().update(fileId=ex["id"], media_body=media).execute()
    else:
        drive.files().create(body={"name": nombre, "parents": [folder_id]},
                             media_body=media, fields="id").execute()

def mover_archivo(file_id, nuevo_parent):
    """Mueve un .txt ya procesado para dejar limpio el input."""
    meta = drive.files().get(fileId=file_id, fields="parents").execute()
    drive.files().update(fileId=file_id, addParents=nuevo_parent,
                         removeParents=",".join(meta.get("parents", [])), fields="id").execute()

print("Funciones de Drive listas")
#cuarto

def _norm(s):
    return unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode().lower().strip()

def _col(df, nombre):
    """Encuentra la columna aunque cambien tildes o mayúsculas"""
    obj = _norm(nombre)
    for c in df.columns:
        if _norm(c) == obj:
            return c
    raise KeyError(f"No encuentro la columna '{nombre}'. Disponibles: {list(df.columns)}")

def _a_numero(serie):
    """Convierte texto a número: tolera '1.234,56', '1234.56', '-', vacíos"""
    def conv(x):
        x = str(x).strip().replace(" ", "")
        if x in ("", "-", "nan", "na", "none", "NA", "None"):
            return 0.0
        p, c = "." in x, "," in x
        if p and c:   x = x.replace(".", "").replace(",", ".")   # 1.234.567,89
        elif c:       x = x.replace(",", ".")                    # 1234,56
        try:    return float(x)
        except: return 0.0
    return serie.map(conv)

def leer_txt(ruta):
    try:    return pd.read_csv(ruta, sep=SEPARADOR, dtype=str, encoding="utf-8")
    except UnicodeDecodeError:
            return pd.read_csv(ruta, sep=SEPARADOR, dtype=str, encoding="latin-1")

print("Funciones de cálculo listas")



#quinto

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timezone, timedelta

def _hora_bogota(): return datetime.now(timezone(timedelta(hours=-5)))
def _sheet_name(x):
    n = re.sub(r'[\\/*?:\[\]]', "_", str(x)).strip()[:31]
    return n or "SIN_NOMBRE"
def _v(x):  # numpy -> python nativo para openpyxl
    return x.item() if hasattr(x, "item") else x
def _cols_txt(df):  # columnas originales del TXT (sin las internas _deb/_cre)
    return [c for c in df.columns if not c.startswith("_")]
def _bloque(xw, sheet, df, startrow, gap=2):
    """Escribe un DataFrame y devuelve la fila (0-based) libre siguiente."""
    df.to_excel(xw, sheet_name=sheet, startrow=startrow, index=False)
    return startrow + len(df) + 1 + gap

# ---------- PASO 1: reporte general (hojas por comprobante) ----------
def escribir_hoja_comprobante(xw, sheet, sub, titulo, ts, cols):
    piv = (sub.groupby([cols["cuenta"], cols["tipo"], cols["desc"]], dropna=False)
              .agg(**{"Suma de Debito":  ("_deb", "sum"),
                      "Suma de Crédito": ("_cre", "sum"),
                      "Suma de Neto":    ("Neto", "sum")})
              .reset_index()
              .rename(columns={cols["cuenta"]: "Cuenta", cols["tipo"]: "Tipo Comprobante",
                               cols["desc"]: "Descripción transacción"}))
    n4 = sub.loc[sub["Extrae"] == "4", "Neto"].sum()
    n2 = sub.loc[sub["Extrae"] == "2", "Neto"].sum(); dif = n4 + n2
    cuadre = pd.DataFrame({"Concepto": ["Neto cuentas 4", "Neto cuentas 2", "DIFERENCIA (4+2)"],
                           "Valor": [n4, n2, dif]})
    row = 3
    row = _bloque(xw, sheet, piv, row)                 # tabla dinámica
    row = _bloque(xw, sheet, cuadre, row)              # cuadro de sumas
    ws = xw.sheets[sheet]
    ws.cell(row + 1, 1, "DETALLE (líneas del TXT)")    # rótulo
    _bloque(xw, sheet, sub[_cols_txt(sub)], row + 1, gap=0)   # líneas crudas del TXT, al final
    ws.cell(1, 1, titulo); ws.cell(2, 1, f"Generado: {ts}")
    return {"Comprobante": sheet, "Neto_4": n4, "Neto_2": n2, "Diferencia": dif,
            "Estado": "OK" if abs(dif) <= TOLERANCIA else ">>> REVISAR"}

def escribir_resumen_general(xw, filas, ts, detalle_txt):
    row = 2
    row = _bloque(xw, "RESUMEN_GENERAL", pd.DataFrame(filas), row)
    ws = xw.sheets["RESUMEN_GENERAL"]
    ws.cell(1, 1, f"RESUMEN GENERAL — Generado: {ts}")
    ws.cell(row + 1, 1, "DETALLE (todas las líneas del TXT)")
    _bloque(xw, "RESUMEN_GENERAL", detalle_txt, row + 1, gap=0)  # TODO el TXT del lote, al final

# ---------- PASO 2: hoja por cuenta IVA dentro del Excel del comprobante ----------
def escribir_hoja_iva_ws(ws, sub, cuenta_iva, titulo, ts, cols):
    ing = sub[sub["Extrae"] == "4"]
    piv = (ing.groupby([cols["cuenta"], cols["desc"]], dropna=False)
              .agg(**{"Suma Débito":  ("_deb", "sum"),
                      "Suma Crédito": ("_cre", "sum"),
                      "Suma Neto":    ("Neto", "sum")})
              .reset_index()
              .rename(columns={cols["cuenta"]: "Cuenta Ingreso", cols["desc"]: "Descripción Transacción"}))
    piv.insert(0, "Cuenta IVA", cuenta_iva)
    piv = piv[["Cuenta IVA", "Cuenta Ingreso", "Descripción Transacción",
               "Suma Débito", "Suma Crédito", "Suma Neto"]]
    n4 = float(sub.loc[sub["Extrae"] == "4", "Neto"].sum())
    n2 = float(sub.loc[sub["Extrae"] == "2", "Neto"].sum()); dif = n4 + n2
    ws.cell(1, 1, titulo); ws.cell(2, 1, f"Generado: {ts}")
    fila = 4
    for r in dataframe_to_rows(piv, index=False, header=True):       # tabla dinámica
        for j, val in enumerate(r, start=1): ws.cell(fila, j, _v(val))
        fila += 1
    fila += 1
    for concepto, valor in [("Concepto", "Valor"), ("Neto cuentas 4", n4),   # cuadro de sumas
                            ("Neto cuentas 2", n2), ("DIFERENCIA (4+2)", dif)]:
        ws.cell(fila, 1, concepto); ws.cell(fila, 2, valor); fila += 1
    fila += 1
    ws.cell(fila, 1, "DETALLE (líneas del TXT)"); fila += 1                    # rótulo
    for r in dataframe_to_rows(sub[_cols_txt(sub)], index=False, header=True):  # líneas crudas, al final
        for j, val in enumerate(r, start=1): ws.cell(fila, j, _v(val))
        fila += 1

print("✅ Funciones de Paso 1 y Paso 2 listas (con detalle del TXT al final)")

#sexto
assert FOLDER_ID.strip(), "Falta el FOLDER_ID en el Bloque 2"

# Las carpetas de salida se crean AL MISMO NIVEL que tu carpeta de TXT (no adentro)
padre = drive.files().get(fileId=FOLDER_ID, fields="parents").execute().get("parents", ["root"])[0]
f_rep  = obtener_o_crear_subcarpeta(SUB_REPORTE, padre)
f_comp = obtener_o_crear_subcarpeta(SUB_COMPROBANTES, padre)
f_proc = obtener_o_crear_subcarpeta(SUB_TXT_PROC, padre)

archivos = listar_archivos(FOLDER_ID)
assert archivos, " No hay .txt/.csv en la carpeta. Sube el lote y vuelve a correr."
print(f" Lote actual: {len(archivos)} archivo(s)")

# 1) Consolidar
frames = []
for f in archivos:
    local = descargar(f["id"], f"/content/{f['name']}")
    d = leer_txt(local); d.columns = d.columns.str.strip(); frames.append(d)
master = pd.concat(frames, ignore_index=True)

# 2) Extrae y Neto
cC = _col(master, COL_CUENTA);  cD = _col(master, COL_DEBITO)
cR = _col(master, COL_CREDITO); cT = _col(master, COL_TIPO); cX = _col(master, COL_DESCTRX)
deb, cre = _a_numero(master[cD]), _a_numero(master[cR])
master["Extrae"] = master[cC].astype(str).str.strip().str[0]
master["Neto"] = np.select(
    [master["Extrae"] == "4", master["Extrae"] == "2"],
    [(deb - cre) * IVA, (cre - deb) * SIGNO_CUENTA_2], default=0.0)
master["_deb"], master["_cre"] = deb, cre
cols = {"cuenta": cC, "tipo": cT, "desc": cX}

# 3) Cuenta IVA = la ÚNICA que empieza por 2
cuentas2 = sorted(master.loc[master["Extrae"] == "2", cC].astype(str).str.strip().unique())
assert len(cuentas2) != 0, " El lote NO tiene cuenta que empiece por 2."
assert len(cuentas2) == 1, f" Hay más de una cuenta que empieza por 2: {cuentas2}. Sube un lote por caso."
cuenta_iva = cuentas2[0]
print(f" Cuenta IVA del lote: {cuenta_iva}")

ts    = _hora_bogota().strftime("%d/%m/%Y %H:%M:%S")
sello = _hora_bogota().strftime("%Y-%m-%d_%H%M%S")

ruta_rep = f"/content/{cuenta_iva}_{sello}.xlsx"; filas = []
with pd.ExcelWriter(ruta_rep, engine="openpyxl") as xw:
    pd.DataFrame().to_excel(xw, sheet_name="RESUMEN_GENERAL", index=False)
    for tipo, gt in master.groupby(cT):
        filas.append(escribir_hoja_comprobante(xw, _sheet_name(tipo), gt,
                          f"COMPROBANTE: {tipo}  |  CUENTA IVA: {cuenta_iva}", ts, cols))
    escribir_resumen_general(xw, filas, ts, master[_cols_txt(master)])
subir(ruta_rep, f_rep)
print(f" Paso 1 listo: {cuenta_iva}_{sello}.xlsx")

for tipo, gt in master.groupby(cT):
    nombre = f"{_sheet_name(tipo)}.xlsx"; local = f"/content/{nombre}"
    ex = buscar_archivo(nombre, f_comp)
    if ex:                                   # ya existe -> lo bajo y le agrego hoja
        descargar(ex["id"], local); wb = openpyxl.load_workbook(local)
    else:                                    # no existe -> libro nuevo
        wb = openpyxl.Workbook(); wb.remove(wb.active)
    sh = _sheet_name(cuenta_iva)
    if sh in wb.sheetnames: del wb[sh]       # si esa cuenta ya estaba, la reemplazo
    ws = wb.create_sheet(title=sh)
    escribir_hoja_iva_ws(ws, gt, cuenta_iva,
                         f"CUENTA IVA: {cuenta_iva}  |  COMPROBANTE: {tipo}", ts, cols)
    wb.save(local)
    actualizar_o_crear(nombre, f_comp, local)
print(f" Paso 2 listo: hoja '{cuenta_iva}' agregada en {master[cT].nunique()} comprobante(s)")

for f in archivos:
    mover_archivo(f["id"], f_proc)
print(" Lote terminado. Sube el siguiente y vuelve a correr el Bloque 6.")




