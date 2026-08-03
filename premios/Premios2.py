# Bloque 1
!pip install xlsxwriter -q

import pandas as pd
import numpy as np
import glob
import os
import re
import shutil
import time
import unicodedata
from datetime import datetime, timedelta

from google.colab import drive
drive.mount('/content/drive')

print("Librerías cargadas.")

#Bloque 2
# === Rutas ===
CARPETA_BASE       = "/content/drive/My Drive/Optimizacion-Premios/2026/"
CARPETA_ENTRADA    = os.path.join(CARPETA_BASE, "Entrada")
CARPETA_SALIDA     = os.path.join(CARPETA_BASE, "Salida")
CARPETA_PROCESADOS = os.path.join(CARPETA_BASE, "Procesados")

ANIO_GRAVABLE = 2026

# === Control del movimiento de archivos ===
# Déjalo en False mientras pruebas. Cuando confirmes que el consolidado
# sale bien, cámbialo a True y corre el BLOQUE 10.
MOVER_PROCESADOS = False

# === Estructura del archivo ===
FILA_ENCABEZADO = None      # None = detectar. Si falla: 6 (fila de Excel)

# === Columnas clave ===
COL_VALOR  = None           # None = buscar por encabezado. Si falla: "M"
COL_PREMIO = None           # None = buscar por encabezado. Si falla: "K"
COL_MES    = None           # None = buscar por encabezado. Si falla: "B"

# === Tarifas y topes ===
UVT = 52_374                # <-- VERIFICAR el UVT del año gravable
TOPE_PREMIOS   = 48 * UVT
TOPE_OTROS     = 27 * UVT
TARIFA_PREMIOS = 20.0
TARIFA_OTROS   = 3.5

for c in (CARPETA_ENTRADA, CARPETA_SALIDA, CARPETA_PROCESADOS):
    os.makedirs(c, exist_ok=True)

print(f"Entrada:    {CARPETA_ENTRADA}")
print(f"Salida:     {CARPETA_SALIDA}")
print(f"Procesados: {CARPETA_PROCESADOS}")
print(f"\nMover procesados: {MOVER_PROCESADOS}")
print(f"Tope premios (48 UVT): {TOPE_PREMIOS:,.0f}")
print(f"Tope otros   (27 UVT): {TOPE_OTROS:,.0f}")

#Bloque 3 
def sin_tildes(texto):
    return "".join(c for c in unicodedata.normalize("NFD", str(texto))
                   if unicodedata.category(c) != "Mn")

def limpiar_encabezado(texto):
    return re.sub(r"\s+", " ", sin_tildes(texto).upper().strip())

def letra_a_indice(letra):
    idx = 0
    for c in str(letra).strip().upper():
        idx = idx * 26 + (ord(c) - 64)
    return idx - 1

def indice_a_letra(i):
    s = ""; i += 1
    while i:
        i, r = divmod(i - 1, 26)
        s = chr(65 + r) + s
    return s


# ---------------------------------------------------------------- NÚMEROS
def normalizar_numero(valor):
    """1.234.567,89 | 1,234,567.89 | $ 1.234.567 | (1.234) | 1234"""
    if isinstance(valor, (int, float, np.integer, np.floating)):
        return np.nan if pd.isna(valor) else float(valor)
    if valor is None:
        return np.nan

    s = str(valor).replace("\xa0", " ")
    s = s.replace("\u2212", "-").replace("–", "-")
    s = re.sub(r"[^\d,\.\-()]", "", s).strip()
    if s in ("", "-", ".", ","):
        return np.nan

    negativo = (s.startswith("(") and s.endswith(")")) or s.startswith("-")
    s = s.strip("()").lstrip("-")

    hay_coma, hay_punto = "," in s, "." in s
    if hay_coma and hay_punto:
        # el separador MÁS A LA DERECHA es el decimal
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")   # formato colombiano
        else:
            s = s.replace(",", "")                     # formato americano
    elif hay_coma:
        p = s.split(",")
        s = s.replace(",", ".") if (len(p) == 2 and len(p[1]) != 3) else s.replace(",", "")
    elif hay_punto:
        p = s.split(".")
        if not (len(p) == 2 and len(p[1]) != 3):
            s = s.replace(".", "")

    try:
        v = float(s)
    except ValueError:
        return np.nan
    return -v if negativo else v


# ------------------------------------------------------------------ SI/NO
_SI = {"SI", "S", "1", "1.0", "X", "TRUE", "VERDADERO", "SIP", "SII"}
_NO = {"NO", "N", "0", "0.0", "FALSE", "FALSO", "NA", "N/A", "NINGUNO"}

def normalizar_si_no(valor):
    """si / Si / SÍ / sí. / 1 / x -> 'SI'    |    no / No / N -> 'NO'"""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return ""
    s = sin_tildes(valor).upper().strip().replace(".", "").replace(" ", "")
    if s in _SI:
        return "SI"
    if s in _NO:
        return "NO"
    return s


# ------------------------------------------------------------------- MESES
MESES_NOMBRE = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
                5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
                9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}

_ALIAS_MES = {}
for _n, _nom in MESES_NOMBRE.items():
    _b = sin_tildes(_nom).upper()
    _ALIAS_MES[_b] = _n
    _ALIAS_MES[_b[:3]] = _n
    _ALIAS_MES[f"{_n:02d}"] = _n
_ALIAS_MES.update({
    "SETIEMBRE": 9, "SET": 9, "SEPT": 9, "SBRE": 9, "SEPTBRE": 9,
    "AGTO": 8, "AGST": 8, "OCTB": 10, "OCTBRE": 10,
    "NVBRE": 11, "NOVBRE": 11, "DICBRE": 12, "DBRE": 12,
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4, "JUNE": 6,
    "JULY": 7, "AUGUST": 8, "SEPTEMBER": 9, "OCTOBER": 10,
    "NOVEMBER": 11, "DECEMBER": 12,
})

def normalizar_mes(valor):
    """enero / ENERO / Ene. / Énero / 1 / 01 / 'ENERO 2026' / fecha real"""
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return (None, "")
    if isinstance(valor, (datetime, pd.Timestamp)):
        return (valor.month, MESES_NOMBRE[valor.month])
    if isinstance(valor, (int, float, np.integer, np.floating)):
        if pd.isna(valor):
            return (None, "")
        n = int(valor)
        return (n, MESES_NOMBRE[n]) if 1 <= n <= 12 else (None, "")

    s = sin_tildes(valor).upper().strip().replace(".", "").replace("_", " ")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return (None, "")
    if s in _ALIAS_MES:
        n = _ALIAS_MES[s]
        return (n, MESES_NOMBRE[n])
    for token in re.split(r"[\s/\-]+", s):
        if token in _ALIAS_MES:
            n = _ALIAS_MES[token]
            return (n, MESES_NOMBRE[n])
    try:
        f = pd.to_datetime(s, dayfirst=True)
        return (f.month, MESES_NOMBRE[f.month])
    except Exception:
        return (None, "")


# --------------------------------------------------- BÚSQUEDA DE COLUMNAS
def buscar_columna(columnas, exactos=(), contiene=(), excluir=()):
    limpias = [limpiar_encabezado(c) for c in columnas]
    for i, h in enumerate(limpias):
        if h in exactos and not any(x in h for x in excluir):
            return i
    for i, h in enumerate(limpias):
        if any(t in h for t in contiene) and not any(x in h for x in excluir):
            return i
    return None


def detectar_fila_encabezado(crudo, max_filas=25):
    """La fila de encabezado es la primera con mayoría de celdas de texto."""
    mejor, mejor_score = None, 0
    for i in range(min(max_filas, len(crudo))):
        fila = crudo.iloc[i]
        textos = sum(1 for v in fila if isinstance(v, str) and v.strip())
        llenas = fila.notna().sum()
        if llenas and textos / max(llenas, 1) > 0.7 and textos > mejor_score:
            mejor, mejor_score = i, textos
    return mejor


# --- prueba rápida ---
print("SI/NO:")
for p in ["sí", "SI ", "no", "N", "x", "1", "Talvez"]:
    print(f"   {p!r:12} -> {normalizar_si_no(p)!r}")
print("\nMESES:")
for m in ["enero", "ENERO", "Ene.", "01", 3, "SETIEMBRE", "MARZO 2026", "Trimestre 1"]:
    print(f"   {m!r:15} -> {normalizar_mes(m)}")
print("\nNÚMEROS:")
for n in ["1.234.567,89", "1,234,567.89", "$ 2.514.000", "(1.234)", 15000, "-"]:
    print(f"   {n!r:16} -> {normalizar_numero(n)}")

#Bloque 4 
EXTENSIONES = (".xlsx", ".xlsm", ".xls")

if not os.path.isdir(CARPETA_ENTRADA):
    raise FileNotFoundError(f"La carpeta no existe: {CARPETA_ENTRADA}")

archivos, ignorados = [], []
for f in sorted(os.listdir(CARPETA_ENTRADA)):
    ruta = os.path.join(CARPETA_ENTRADA, f)
    if not os.path.isfile(ruta) or f.startswith(("~$", ".")):
        continue
    if os.path.splitext(f)[1].lower() in EXTENSIONES:
        archivos.append(ruta)
    else:
        ignorados.append(f)

if ignorados:
    print("[AVISO] Ignorados por extensión:")
    for f in ignorados:
        nota = "  <-- posible Google Sheet nativo" if not os.path.splitext(f)[1] else ""
        print(f"    {f!r}{nota}")
    print()

if not archivos:
    raise FileNotFoundError(f"No hay archivos de Excel en: {CARPETA_ENTRADA}")

print(f"Archivos a procesar: {len(archivos)}\n")

frames = []
columnas_referencia = None
ARCHIVOS_OK = []        # rutas que SÍ entraron al consolidado -> se mueven
ARCHIVOS_FALLA = []     # rutas con problema -> se quedan en Entrada

for archivo in archivos:
    nombre = os.path.basename(archivo)
    try:
        crudo = pd.read_excel(archivo, sheet_name=0, header=None, dtype=object)
    except Exception as e:
        print(f"  [ERROR] {nombre}: {e}")
        ARCHIVOS_FALLA.append((archivo, str(e)))
        continue

    fila_enc = (FILA_ENCABEZADO - 1) if FILA_ENCABEZADO else detectar_fila_encabezado(crudo)
    if fila_enc is None:
        print(f"  [ERROR] {nombre}: no se detectó fila de encabezado.")
        ARCHIVOS_FALLA.append((archivo, "sin fila de encabezado"))
        continue

    df = pd.read_excel(archivo, sheet_name=0, header=fila_enc, dtype=object)
    df = df.dropna(how="all").dropna(axis=1, how="all")
    df = df.loc[:, [c for c in df.columns if not str(c).startswith("Unnamed:")]]

    primera = df.columns[0]
    df = df[~df[primera].astype(str).str.upper().str.strip()
            .isin([limpiar_encabezado(primera), "TOTAL", "TOTALES", "NAN", ""])]

    if df.empty:
        print(f"  [AVISO] {nombre}: quedó vacío tras la limpieza.")
        ARCHIVOS_FALLA.append((archivo, "vacío tras limpieza"))
        continue

    # pd.concat une por NOMBRE de columna: si un archivo trae 'VALOR ' con
    # espacio y otro 'VALOR', pandas crea dos columnas y todo se desalinea.
    if columnas_referencia is None:
        columnas_referencia = list(df.columns)
    elif list(df.columns) != columnas_referencia:
        if len(df.columns) == len(columnas_referencia):
            print(f"  [AVISO] {nombre}: encabezados distintos -> alineado por posición.")
            df.columns = columnas_referencia
        else:
            print(f"  [ERROR] {nombre}: {len(df.columns)} columnas vs "
                  f"{len(columnas_referencia)} del primero. ESTRUCTURA DISTINTA.")
            ARCHIVOS_FALLA.append((archivo, "número de columnas distinto"))
            continue

    df["Archivo Origen"] = nombre
    frames.append(df)
    ARCHIVOS_OK.append(archivo)
    print(f"  OK  {nombre}  |  encabezado fila {fila_enc + 1}  |  "
          f"{len(df)} filas x {len(df.columns) - 1} cols")

if not frames:
    raise ValueError("Ningún archivo pudo procesarse.")

df_consolidado = pd.concat(frames, ignore_index=True)

print(f"\nConsolidado: {len(df_consolidado)} filas x {len(df_consolidado.columns)} columnas")
print(f"Procesados OK: {len(ARCHIVOS_OK)}   |   Con falla: {len(ARCHIVOS_FALLA)}")
if ARCHIVOS_FALLA:
    print("\nEstos se quedan en Entrada para revisión:")
    for r, motivo in ARCHIVOS_FALLA:
        print(f"    {os.path.basename(r)}  ->  {motivo}")

#Bloque 5 
cols = [c for c in df_consolidado.columns if c != "Archivo Origen"]

print("=" * 100)
print(f"{'LETRA':<7}{'IDX':<5}{'ENCABEZADO':<40}{'MUESTRA'}")
print("=" * 100)
for i, c in enumerate(cols):
    m = " | ".join(str(v)[:16] for v in df_consolidado[c].dropna().head(3))
    print(f"{indice_a_letra(i):<7}{i:<5}{str(c)[:38]:<40}{m[:44]}")
print("=" * 100 + "\n")

idx_valor = (letra_a_indice(COL_VALOR) if COL_VALOR else buscar_columna(
    cols,
    exactos=("VALOR", "BASE", "IMPORTE", "MONTO", "VALOR BASE", "BASE GRAVABLE"),
    contiene=("VALOR", "BASE", "IMPORTE", "MONTO", "DEBITO", "CREDITO", "SALDO"),
    excluir=("RETENCION", "TARIFA", "IVA", "FECHA", "CODIGO")))

idx_premio = (letra_a_indice(COL_PREMIO) if COL_PREMIO else buscar_columna(
    cols,
    exactos=("PREMIO", "ES PREMIO", "PREMIOS", "APLICA PREMIO", "SI/NO"),
    contiene=("PREMIO",)))

idx_mes = (letra_a_indice(COL_MES) if COL_MES else buscar_columna(
    cols,
    exactos=("MES", "PERIODO", "PERIODO GRAVABLE", "MES DECLARADO"),
    contiene=("MES", "PERIODO", "FECHA")))

print("COLUMNAS RESUELTAS:")
faltan = []
for etiqueta, idx, param in [("VALOR / BASE", idx_valor,  "COL_VALOR"),
                             ("SI/NO PREMIO", idx_premio, "COL_PREMIO"),
                             ("MES",          idx_mes,    "COL_MES")]:
    if idx is None:
        print(f"   [X] {etiqueta:<14} NO ENCONTRADA -> define {param} en el BLOQUE 2")
        faltan.append(param)
    else:
        print(f"   [OK] {etiqueta:<14} col {indice_a_letra(idx)} (idx {idx}) -> '{cols[idx]}'")
        print(f"        muestra: {df_consolidado[cols[idx]].dropna().head(4).tolist()}")

if faltan:
    raise ValueError(f"Define estos parámetros en el BLOQUE 2 y vuelve a correr: {faltan}")

#Bloque 6
col_valor, col_premio, col_mes = cols[idx_valor], cols[idx_premio], cols[idx_mes]

df_consolidado["Base Normalizada"] = df_consolidado[col_valor].map(normalizar_numero)
df_consolidado["Es Premio"]        = df_consolidado[col_premio].map(normalizar_si_no)

_m = df_consolidado[col_mes].map(normalizar_mes)
df_consolidado["Mes Num"] = [x[0] for x in _m]
df_consolidado["Mes"]     = [x[1] for x in _m]
df_consolidado["Periodo"] = [f"{ANIO_GRAVABLE}-{int(n):02d}" if pd.notna(n) else ""
                             for n in df_consolidado["Mes Num"]]

print("--- CONTROL DE CALIDAD ---\n")

n = df_consolidado["Base Normalizada"].isna().sum()
if n:
    print(f"[!] {n} filas con base no numérica. Valores crudos:")
    print("   ", df_consolidado.loc[df_consolidado['Base Normalizada'].isna(), col_valor]
          .dropna().astype(str).unique()[:10], "\n")
else:
    print("[OK] Todas las bases se convirtieron a número.\n")

raros = sorted(set(df_consolidado["Es Premio"]) - {"SI", "NO", ""})
if raros:
    print(f"[!] Valores no reconocidos en '{col_premio}': {raros}\n")
else:
    print("[OK] Columna SI/NO normalizada sin residuos.\n")

n = df_consolidado["Mes Num"].isna().sum()
if n:
    print(f"[!] {n} filas sin mes reconocible. Valores crudos:")
    print("   ", df_consolidado.loc[df_consolidado['Mes Num'].isna(), col_mes]
          .dropna().astype(str).unique()[:10], "\n")
else:
    print("[OK] Todos los meses se reconocieron.\n")

print("Meses detectados:")
print(df_consolidado["Mes"].value_counts().to_string())
print(f"\nSuma total de bases: {df_consolidado['Base Normalizada'].sum():,.0f}")

#Bloque 7
df_consolidado["Tarifa Retención (%)"] = 0.0

es_premio = df_consolidado["Es Premio"].eq("SI")
no_premio = df_consolidado["Es Premio"].eq("NO")
base      = df_consolidado["Base Normalizada"]

df_consolidado.loc[es_premio & base.gt(TOPE_PREMIOS), "Tarifa Retención (%)"] = TARIFA_PREMIOS
df_consolidado.loc[no_premio & base.gt(TOPE_OTROS),   "Tarifa Retención (%)"] = TARIFA_OTROS

df_consolidado["Valor Retención"] = (base * df_consolidado["Tarifa Retención (%)"] / 100).round(0)
df_consolidado.loc[base.isna(), "Valor Retención"] = np.nan

print("--- RESUMEN ---")
print(df_consolidado.groupby("Tarifa Retención (%)")
      .agg(Filas=("Base Normalizada", "size"),
           Base=("Base Normalizada", "sum"),
           Retencion=("Valor Retención", "sum"))
      .to_string(float_format=lambda x: f"{x:,.0f}"))

#Bloque 8

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def dar_formato(ruta):
    wb = openpyxl.load_workbook(ruta)
    f_enc    = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    relleno  = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    centro   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    f_cuerpo = Font(name="Calibri", size=11)

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        if ws.max_row < 1:
            continue
        enc = [c.value for c in ws[1]]

        for col in ws.columns:
            L = get_column_letter(col[0].column)
            # el str() faltaba en la versión original: len() explotaba en
            # celdas numéricas y el except vacío se lo tragaba
            largo = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[L].width = min(max(largo + 2, 10), 40)

        for c in ws[1]:
            c.font, c.fill, c.alignment = f_enc, relleno, centro

        for fila in ws.iter_rows(min_row=2):
            for c in fila:
                c.alignment, c.font = centro, f_cuerpo
                if isinstance(c.value, (int, float)):
                    t = str(enc[c.column - 1]) if c.column - 1 < len(enc) else ""
                    c.number_format = ("0.0" if "Tarifa" in t
                                       else "0" if "Mes Num" in t else "#,##0")
        ws.freeze_panes = "A2"

    wb.save(ruta)
    print(f"Formateado: {os.path.basename(ruta)}")

print("Función dar_formato() lista.")

#Bloque 9
# --- imports locales: funciona aunque se haya reiniciado el entorno ---
import os, shutil, time
import pandas as pd, numpy as np
from datetime import datetime, timedelta

requeridas = ["df_consolidado", "CARPETA_SALIDA", "ANIO_GRAVABLE",
              "MESES_NOMBRE", "TARIFA_PREMIOS", "TARIFA_OTROS"]
faltantes = [v for v in requeridas if v not in globals()]
if faltantes:
    raise NameError(f"Faltan variables: {faltantes}\n"
                    f"Se reinició el entorno. Corre los BLOQUES 1 a 7 en orden.")

os.makedirs(CARPETA_SALIDA, exist_ok=True)

base      = df_consolidado["Base Normalizada"]
es_premio = df_consolidado["Es Premio"].eq("SI")
no_premio = df_consolidado["Es Premio"].eq("NO")

validos = df_consolidado["Mes Num"].dropna()
if len(validos):
    mes_num = int(validos.mode().iloc[0])          # el punto que faltaba
    if validos.nunique() > 1:
        print(f"[AVISO] Hay {validos.nunique()} meses en la entrada. "
              f"Se nombra con el predominante: {MESES_NOMBRE[mes_num]}")
else:
    ant = datetime.now().replace(day=1) - timedelta(days=1)
    mes_num = ant.month
    print("[AVISO] Ningún mes reconocido. Se usa el mes anterior calendario.")

nombre_mes  = MESES_NOMBRE[mes_num]
ETIQUETA_MES = f"{mes_num:02d}_{nombre_mes.lower()}"
ruta_salida = os.path.join(
    CARPETA_SALIDA, f"Consolidado_con_Retencion-{ETIQUETA_MES}_{ANIO_GRAVABLE}.xlsx")

sin_clasificar = (base.isna()
                  | ~df_consolidado["Es Premio"].isin(["SI", "NO"])
                  | df_consolidado["Mes Num"].isna())

hojas = {
    "Consolidado General":            df_consolidado,
    "Premios20%":                     df_consolidado[df_consolidado["Tarifa Retención (%)"] == TARIFA_PREMIOS],
    "Otros ingresos tributarios3.5%": df_consolidado[df_consolidado["Tarifa Retención (%)"] == TARIFA_OTROS],
    "Premios Sin Retencion":          df_consolidado[es_premio & df_consolidado["Tarifa Retención (%)"].eq(0) & base.notna()],
    "Otros Ingresos Sin retencion":   df_consolidado[no_premio & df_consolidado["Tarifa Retención (%)"].eq(0) & base.notna()],
    "REVISAR":                        df_consolidado[sin_clasificar],
}

# --- respaldo si ya existía ---
if os.path.exists(ruta_salida):
    hist = os.path.join(CARPETA_SALIDA, "Historico")
    os.makedirs(hist, exist_ok=True)
    b, e = os.path.splitext(os.path.basename(ruta_salida))
    shutil.copy2(ruta_salida, os.path.join(hist, f"{b}__{datetime.now():%Y%m%d_%H%M%S}{e}"))
    print("[BACKUP] Versión anterior copiada a Historico/\n")

with pd.ExcelWriter(ruta_salida, engine="xlsxwriter") as w:
    for hoja, data in hojas.items():
        data.to_excel(w, sheet_name=hoja[:31], index=False)
        print(f"  Hoja '{hoja}': {len(data)} filas")

if "dar_formato" in globals():
    dar_formato(ruta_salida)
else:
    print("[AVISO] 'dar_formato' no está definida. Corre el BLOQUE 8 y repite.")

# --- confirmar escritura en Drive (tiene retardo de sincronización) ---
EXPORTACION_OK = False
for _ in range(15):
    if os.path.exists(ruta_salida) and os.path.getsize(ruta_salida) > 0:
        EXPORTACION_OK = True
        print(f"\n[OK] Guardado: {ruta_salida}")
        print(f"     {os.path.getsize(ruta_salida):,} bytes")
        break
    time.sleep(1)

if not EXPORTACION_OK:
    print(f"\n[X] NO se confirmó la escritura. NO muevas los archivos todavía.")

#Bloque 10
if not globals().get("EXPORTACION_OK", False):
    raise RuntimeError("La exportación no se confirmó. Revisa el BLOQUE 9 "
                       "antes de mover nada.")

if not MOVER_PROCESADOS:
    print("MOVER_PROCESADOS = False  ->  no se movió nada.\n")
    print(f"Archivos listos para mover ({len(ARCHIVOS_OK)}):")
    for r in ARCHIVOS_OK:
        print(f"    {os.path.basename(r)}")
    print(f"\nDestino: {os.path.join(CARPETA_PROCESADOS, str(ANIO_GRAVABLE), ETIQUETA_MES)}")
    print("\nCambia MOVER_PROCESADOS = True en el BLOQUE 2 y vuelve a correr este bloque.")

else:
    destino = os.path.join(CARPETA_PROCESADOS, str(ANIO_GRAVABLE), ETIQUETA_MES)
    os.makedirs(destino, exist_ok=True)

    movidos, errores = [], []
    for origen in ARCHIVOS_OK:
        nombre = os.path.basename(origen)

        if not os.path.exists(origen):
            print(f"  [SKIP] {nombre}: ya no está en Entrada.")
            continue

        ruta_destino = os.path.join(destino, nombre)
        # Si ya existe uno con el mismo nombre, no lo pisamos
        if os.path.exists(ruta_destino):
            b, e = os.path.splitext(nombre)
            ruta_destino = os.path.join(destino, f"{b}__{datetime.now():%Y%m%d_%H%M%S}{e}")
            print(f"  [DUP] Ya existía {nombre} -> se guarda como "
                  f"{os.path.basename(ruta_destino)}")

        try:
            # copy2 + remove en vez de move: en Drive montado, move puede
            # fallar a mitad y dejar el archivo perdido. Así, si la copia
            # falla, el original sigue intacto en Entrada.
            shutil.copy2(origen, ruta_destino)
            if os.path.exists(ruta_destino) and os.path.getsize(ruta_destino) > 0:
                os.remove(origen)
                movidos.append(nombre)
                print(f"  OK  {nombre}")
            else:
                errores.append((nombre, "la copia quedó vacía"))
                print(f"  [X] {nombre}: la copia quedó vacía, NO se borró el original.")
        except Exception as ex:
            errores.append((nombre, str(ex)))
            print(f"  [X] {nombre}: {ex}")

    # --- bitácora ---
    log = os.path.join(CARPETA_PROCESADOS, "bitacora_procesamiento.csv")
    registro = pd.DataFrame({
        "Fecha_Proceso":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Anio":            ANIO_GRAVABLE,
        "Mes":             nombre_mes,
        "Archivo":         movidos,
        "Carpeta_Destino": destino,
        "Consolidado":     os.path.basename(ruta_salida),
    })
    if os.path.exists(log):
        registro.to_csv(log, mode="a", header=False, index=False, encoding="utf-8-sig")
    else:
        registro.to_csv(log, index=False, encoding="utf-8-sig")

    print(f"\n{'=' * 60}")
    print(f"Movidos: {len(movidos)}   |   Con error: {len(errores)}")
    print(f"Destino: {destino}")
    print(f"Bitácora: {log}")
    if ARCHIVOS_FALLA:
        print(f"\nSiguen en Entrada para revisión ({len(ARCHIVOS_FALLA)}):")
        for r, motivo in ARCHIVOS_FALLA:
            print(f"    {os.path.basename(r)}  ->  {motivo}")
    print("=" * 60)

#Bloque 11
patron = os.path.join(CARPETA_SALIDA, "Consolidado_con_Retencion-*.xlsx")
archivos_mes = [a for a in sorted(glob.glob(patron))
                if "ANUAL" not in os.path.basename(a).upper()
                and not os.path.basename(a).startswith("~$")]

if not archivos_mes:
    raise FileNotFoundError(f"No hay consolidados mensuales en {CARPETA_SALIDA}")

print(f"Consolidados encontrados: {len(archivos_mes)}\n")

partes = []
for a in archivos_mes:
    try:
        d = pd.read_excel(a, sheet_name="Consolidado General")
        partes.append(d)
        print(f"  OK  {os.path.basename(a)}  ->  {len(d)} filas")
    except Exception as e:
        print(f"  [ERROR] {os.path.basename(a)}: {e}")

df_anual = pd.concat(partes, ignore_index=True)

antes = len(df_anual)
df_anual = df_anual.drop_duplicates()
if antes != len(df_anual):
    print(f"\n[AVISO] Se eliminaron {antes - len(df_anual)} filas duplicadas.")

df_anual = df_anual.sort_values("Mes Num", na_position="last").reset_index(drop=True)
print(f"\nTotal anual: {len(df_anual)} filas")

presentes = set(df_anual["Mes Num"].dropna().astype(int))
faltantes = [MESES_NOMBRE[m] for m in range(1, 13) if m not in presentes]
if faltantes:
    print(f"[AVISO] Meses sin datos: {faltantes}")

resumen_mes = (df_anual.groupby(["Mes Num", "Mes"], dropna=False)
               .agg(Registros=("Base Normalizada", "size"),
                    Base_Total=("Base Normalizada", "sum"),
                    Retencion_Total=("Valor Retención", "sum"))
               .reset_index().sort_values("Mes Num"))

resumen_mes_tarifa = (df_anual.pivot_table(
    index=["Mes Num", "Mes"], columns="Tarifa Retención (%)",
    values=["Base Normalizada", "Valor Retención"],
    aggfunc="sum", fill_value=0).reset_index())
resumen_mes_tarifa.columns = [" ".join(str(x) for x in c if str(x) != "").strip()
                              for c in resumen_mes_tarifa.columns.to_flat_index()]

resumen_tarifa = (df_anual.groupby("Tarifa Retención (%)")
                  .agg(Registros=("Base Normalizada", "size"),
                       Base_Total=("Base Normalizada", "sum"),
                       Retencion_Total=("Valor Retención", "sum")).reset_index())

print("\n--- ANUAL POR MES ---")
print(resumen_mes.to_string(index=False, float_format=lambda x: f"{x:,.0f}"))
print("\n--- ANUAL POR TARIFA ---")
print(resumen_tarifa.to_string(index=False, float_format=lambda x: f"{x:,.0f}"))

ruta_anual = os.path.join(CARPETA_SALIDA, f"CONSOLIDADO_ANUAL_{ANIO_GRAVABLE}.xlsx")

hojas_anual = {
    "Resumen por Mes":                resumen_mes,
    "Resumen Mes x Tarifa":           resumen_mes_tarifa,
    "Resumen por Tarifa":             resumen_tarifa,
    "Consolidado Anual":              df_anual,
    "Premios20%":                     df_anual[df_anual["Tarifa Retención (%)"] == TARIFA_PREMIOS],
    "Otros ingresos tributarios3.5%": df_anual[df_anual["Tarifa Retención (%)"] == TARIFA_OTROS],
    "REVISAR":                        df_anual[df_anual["Base Normalizada"].isna()
                                               | ~df_anual["Es Premio"].isin(["SI", "NO"])
                                               | df_anual["Mes Num"].isna()],
}

with pd.ExcelWriter(ruta_anual, engine="xlsxwriter") as w:
    for hoja, data in hojas_anual.items():
        data.to_excel(w, sheet_name=hoja[:31], index=False)
        print(f"  Hoja '{hoja}': {len(data)} filas")

dar_formato(ruta_anual)
print(f"\nGuardado: {ruta_anual}")

#Bloque 12
def inventario(carpeta, titulo, nivel=0):
    if not os.path.isdir(carpeta):
        print(f"  (no existe: {carpeta})")
        return
    for f in sorted(os.listdir(carpeta)):
        ruta = os.path.join(carpeta, f)
        sangria = "  " * (nivel + 1)
        if os.path.isdir(ruta):
            print(f"{sangria}[{f}]")
            inventario(ruta, titulo, nivel + 1)
        else:
            mod = datetime.fromtimestamp(os.path.getmtime(ruta))
            print(f"{sangria}{f}  ({os.path.getsize(ruta):,} bytes, {mod:%Y-%m-%d %H:%M})")

for carpeta, titulo in [(CARPETA_ENTRADA, "ENTRADA"),
                        (CARPETA_SALIDA, "SALIDA"),
                        (CARPETA_PROCESADOS, "PROCESADOS")]:
    print("=" * 70)
    print(titulo)
    print("=" * 70)
    inventario(carpeta, titulo)
    print()
    
